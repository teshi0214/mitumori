"""
mitumori Agent - Open WebUI Pipe Function
=========================================
- Agent Engine (Reasoning Engine) への接続
- セッション管理（ファイルキャッシュで会話継続）
- Excel出力をOWUI File APIにアップロードしてダウンロードリンクを返す

Valves に設定が必要:
  PROJECT_ID   : GCPプロジェクト番号 (例: 89612432694)
  LOCATION     : リージョン (例: us-central1)
  ENGINE_ID    : Agent Engine ID (例: 2066078006102720512)
  OWUI_BASE_URL: Open WebUIのURL (例: https://your-openwebui.com)
  OWUI_API_KEY : Open WebUI APIキー (管理画面 > アカウント > APIキー)
"""

from pydantic import BaseModel, Field
from typing import Iterator, Union, List, Dict
import requests
import json
import os
import io


CACHE_FILE = "/tmp/owui_session_cache.json"


class Pipe:
    class Valves(BaseModel):
        PROJECT_ID: str = Field(
            default="89612432694",
            description="Google Cloud プロジェクト番号",
        )
        LOCATION: str = Field(
            default="us-central1",
            description="ロケーション",
        )
        ENGINE_ID: str = Field(
            default="2066078006102720512",
            description="Agent Engine ID",
        )
        OWUI_BASE_URL: str = Field(
            default="",
            description="Open WebUI の URL (例: https://your-openwebui.com)",
        )
        OWUI_API_KEY: str = Field(
            default="",
            description="Open WebUI API キー (管理画面 > アカウント > APIキー)",
        )

    def __init__(self):
        self.valves = self.Valves()
        self.name = "見積もり作成 Agent"

    def pipes(self) -> List[Dict[str, str]]:
        return [{"id": "customer_support", "name": "AGENT/Customer Support (Vertex AI Search)"}]

    # ──────────────────────────────────────────
    #  セッションキャッシュ（ファイル永続化）
    # ──────────────────────────────────────────

    def _load_cache(self) -> Dict[str, str]:
        try:
            with open(CACHE_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return {}

    def _save_cache(self, cache: Dict[str, str]):
        try:
            with open(CACHE_FILE, "w") as f:
                json.dump(cache, f)
        except Exception:
            pass

    # ──────────────────────────────────────────
    #  GCP 認証
    # ──────────────────────────────────────────

    def _get_credentials(self):
        from google.auth import default
        from google.auth.transport.requests import Request as GoogleRequest
        credentials, _ = default(scopes=["https://www.googleapis.com/auth/cloud-platform"])
        credentials.refresh(GoogleRequest())
        return credentials

    # ──────────────────────────────────────────
    #  Agent Engine ベースURL
    # ──────────────────────────────────────────

    def _base_url(self):
        return (
            f"https://{self.valves.LOCATION}-aiplatform.googleapis.com/v1/"
            f"projects/{self.valves.PROJECT_ID}/locations/{self.valves.LOCATION}/"
            f"reasoningEngines/{self.valves.ENGINE_ID}"
        )

    # ──────────────────────────────────────────
    #  セッション作成
    # ──────────────────────────────────────────

    def _create_session(self, headers: dict, user_id: str) -> str:
        """Agent Engine 上にセッションを新規作成してsession_idを返す"""
        url = f"{self._base_url()}:query"
        payload = {"class_method": "async_create_session", "input": {"user_id": user_id}}
        resp = requests.post(url=url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        output = data.get("output", {})
        if isinstance(output, str):
            output = json.loads(output)
        session_id = (
            output.get("id")
            or output.get("session_id")
            or output.get("name", "").split("/")[-1]
        )
        print(f"[mitumori] 新規セッション作成: {session_id}")
        return session_id

    # ──────────────────────────────────────────
    #  Excel を OWUI File API にアップロード
    # ──────────────────────────────────────────

    def _upload_excel_to_owui(self, excel_bytes: bytes, filename: str) -> str:
        """
        ExcelバイトデータをOWUI File APIにアップロードし、
        チャット内に表示するダウンロードリンクを返す。
        OWUI_BASE_URL と OWUI_API_KEY が設定されている場合のみ動作。
        """
        if not self.valves.OWUI_BASE_URL or not self.valves.OWUI_API_KEY:
            return None

        try:
            upload_url = f"{self.valves.OWUI_BASE_URL.rstrip('/')}/api/v1/files/"
            headers = {
                "Authorization": f"Bearer {self.valves.OWUI_API_KEY}",
                "Accept": "application/json",
            }
            files = {
                "file": (
                    filename,
                    excel_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            resp = requests.post(upload_url, headers=headers, files=files, timeout=30)
            resp.raise_for_status()
            file_data = resp.json()
            file_id = file_data.get("id")
            if not file_id:
                return None

            download_url = f"{self.valves.OWUI_BASE_URL.rstrip('/')}/api/v1/files/{file_id}/content"
            return download_url
        except Exception as e:
            print(f"[mitumori] OWUI upload error: {e}")
            return None

    # ──────────────────────────────────────────
    #  Agent EngineレスポンスからExcel情報を検出
    # ──────────────────────────────────────────

    def _check_excel_in_response(self, texts: List[str]) -> Dict:
        """
        レスポンステキストにExcel保存成功メッセージがあれば検出する。
        返値: {"detected": bool, "filename": str, "total": str}
        """
        full_text = "\n".join(texts)
        if "を保存しました" in full_text and ".xlsx" in full_text:
            # ファイル名を抽出
            import re
            filename_match = re.search(r"'([^']+\.xlsx)'", full_text)
            filename = filename_match.group(1) if filename_match else "見積もり.xlsx"
            total_match = re.search(r"合計金額: ([0-9,]+)円", full_text)
            total = total_match.group(1) if total_match else ""
            return {"detected": True, "filename": filename, "total": total}
        return {"detected": False}

    # ──────────────────────────────────────────
    #  メインのpipe関数
    # ──────────────────────────────────────────

    def pipe(self, body: dict, __user__: dict = None) -> Union[str, Iterator[str]]:
        try:
            credentials = self._get_credentials()
            headers = {
                "Authorization": f"Bearer {credentials.token}",
                "Content-Type": "application/json",
            }

            # ユーザーメッセージ取得
            messages = body.get("messages", [])
            user_message = ""
            for msg in reversed(messages):
                if msg.get("role") == "user":
                    user_message = msg.get("content", "")
                    break
            if not user_message:
                return "Error: ユーザーメッセージが見つかりません"

            # ユーザーID
            user_id = "default-user"
            if __user__:
                user_id = __user__.get("id", "default-user")

            # chat_id をキーにセッションIDをファイルキャッシュで管理
            chat_id = body.get("chat_id") or body.get("id") or user_id
            cache = self._load_cache()
            if chat_id not in cache:
                session_id = self._create_session(headers, user_id)
                cache[chat_id] = session_id
                self._save_cache(cache)
            else:
                session_id = cache[chat_id]
            print(f"[mitumori] chat_id={chat_id[:8]}... session_id={session_id[:8]}...")

            # streamQuery エンドポイント
            stream_url = f"{self._base_url()}:streamQuery"
            payload = {
                "class_method": "async_stream_query",
                "input": {
                    "message": user_message,
                    "user_id": user_id,
                    "session_id": session_id,
                },
            }

            if body.get("stream", False):
                return self._stream_response(stream_url, payload, headers)

            # 非ストリーミング: 全行受け取ってからExcel処理
            response = requests.post(
                url=stream_url, json=payload, headers=headers,
                stream=True, timeout=120
            )
            response.raise_for_status()

            texts = []
            for line in response.iter_lines():
                if line:
                    try:
                        event = json.loads(line.decode("utf-8"))
                        text = self._extract_text(event)
                        if text:
                            texts.append(text)
                    except json.JSONDecodeError:
                        pass

            result_text = "\n".join(texts) if texts else "応答がありませんでした"

            # Excel保存が検出された場合、OWUIにアップロードしてリンクを追加
            excel_info = self._check_excel_in_response(texts)
            if excel_info["detected"] and self.valves.OWUI_BASE_URL:
                excel_bytes = self._build_excel_from_session(
                    headers, session_id, user_id
                )
                if excel_bytes:
                    filename = excel_info["filename"]
                    download_url = self._upload_excel_to_owui(excel_bytes, filename)
                    if download_url:
                        result_text += (
                            f"\n\n📥 **[{filename} をダウンロード]({download_url})**"
                        )

            return result_text

        except Exception as e:
            import traceback
            return f"Error: {str(e)}\n\n{traceback.format_exc()}"

    # ──────────────────────────────────────────
    #  Agent EngineのstateからExcelを再生成
    # ──────────────────────────────────────────

    def _build_excel_from_session(self, headers: dict, session_id: str, user_id: str) -> bytes:
        """
        Agent EngineのセッションStateからquote:resultを取得してExcelを生成する。
        calculate_quoteの結果がstateに保存されていることが前提。
        """
        try:
            # セッションのstateを取得
            url = f"{self._base_url()}:query"
            payload = {
                "class_method": "async_get_session",
                "input": {"session_id": session_id, "user_id": user_id},
            }
            resp = requests.post(url=url, json=payload, headers=headers, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            output = data.get("output", {})
            if isinstance(output, str):
                output = json.loads(output)

            state = output.get("state", {})
            result = state.get("quote:result")
            if not result:
                return None

            return self._generate_excel_bytes(result)
        except Exception as e:
            print(f"[mitumori] Excel生成エラー: {e}")
            return None

    def _generate_excel_bytes(self, result: dict) -> bytes:
        """quote:result dictからExcelバイトを生成する"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "見積もり"

            thin = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            # タイトル
            ws.merge_cells("A1:G1")
            ws["A1"].value = "御見積書"
            ws["A1"].font = Font(bold=True, size=16)
            ws["A1"].alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 30

            # ヘッダー
            row = 3
            headers_list = ["No.", "品目名", "単価（円）", "数量", "単位", "小計（円）", "備考"]
            widths = [6, 30, 14, 8, 8, 16, 20]
            for col, (h, w) in enumerate(zip(headers_list, widths), 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin
                ws.column_dimensions[get_column_letter(col)].width = w
            row += 1

            # 品目
            for it in result.get("rows", []):
                data = [it["id"], it["name"], it["unit_price"],
                        it["quantity"], it["unit"], it["subtotal"], it.get("note", "")]
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin
                    if col in (3, 6):
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right")
                row += 1

            # 集計
            row += 1
            for label, amount in [
                ("小計", result["subtotal"]),
                (f"消費税（{int(result['tax_rate']*100)}%）", result["tax_amount"]),
            ]:
                ws.merge_cells(f"A{row}:E{row}")
                ws[f"A{row}"].value = label
                ws[f"A{row}"].alignment = Alignment(horizontal="right")
                ws[f"A{row}"].border = thin
                ws[f"F{row}"].value = amount
                ws[f"F{row}"].number_format = '#,##0'
                ws[f"F{row}"].alignment = Alignment(horizontal="right")
                ws[f"F{row}"].border = thin
                row += 1

            # 合計
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"].value = "合計金額"
            ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="right")
            ws[f"A{row}"].border = thin
            ws[f"F{row}"].value = result["total"]
            ws[f"F{row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"F{row}"].fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            ws[f"F{row}"].number_format = '#,##0'
            ws[f"F{row}"].alignment = Alignment(horizontal="right")
            ws[f"F{row}"].border = thin

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except Exception as e:
            print(f"[mitumori] openpyxl error: {e}")
            return None

    # ──────────────────────────────────────────
    #  テキスト抽出
    # ──────────────────────────────────────────

    def _extract_text(self, result) -> str:
        if isinstance(result, dict):
            if "content" in result and "parts" in result["content"]:
                parts = result["content"]["parts"]
                if parts and "text" in parts[0]:
                    return parts[0]["text"]
            elif "text" in result:
                return result["text"]
            elif "output" in result:
                output = result["output"]
                if isinstance(output, str):
                    return output
                elif isinstance(output, dict):
                    return self._extract_text(output)
        return ""

    # ──────────────────────────────────────────
    #  ストリーミングレスポンス
    # ──────────────────────────────────────────

    def _stream_response(self, url: str, payload: dict, headers: dict) -> Iterator[str]:
        def stream_generator():
            try:
                response = requests.post(
                    url=url, json=payload, headers=headers,
                    stream=True, timeout=120
                )
                response.raise_for_status()
                for line in response.iter_lines():
                    if line:
                        try:
                            event = json.loads(line.decode("utf-8"))
                            text = self._extract_text(event)
                            if text:
                                yield text
                        except json.JSONDecodeError:
                            pass
            except Exception as e:
                yield f"\n\nError: {str(e)}"
        return stream_generator()
