"""
見積もりエージェント ツール群
- add_item         : 見積もり品目を追加
- remove_item      : 品目を削除
- list_items       : 現在の品目一覧を表示
- calculate_quote  : Code Execution で小計・割引・税計算を実行
- export_to_excel  : 計算結果を Excel Artifact として保存
- reset_quote      : 見積もりをリセット

ツール間データ受け渡し:
  add/remove/list が tool_context.state["quote:items"] に保存
  calculate_quote が同キーから読み取り計算し state["quote:result"] に保存
  export_to_excel が state["quote:result"] から Excel を生成
"""

import io
import os
from typing import Any

from google.adk.tools import FunctionTool, ToolContext
import google.genai.types as types
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STATE_ITEMS = "quote:items"
STATE_RESULT = "quote:result"


# ────────────────────────────────────────────
#  品目管理ツール
# ────────────────────────────────────────────

def add_item(
    name: str,
    unit_price: float,
    quantity: float,
    unit: str = "式",
    note: str = "",
    tool_context: ToolContext = None,
) -> str:
    """
    見積もりに品目を追加する。

    Args:
        name: 品目名（例: "クラウド設計費"）
        unit_price: 単価（円）
        quantity: 数量
        unit: 単位（例: "式", "時間", "本"）
        note: 備考（オプション）
    """
    items: list[dict] = []
    if tool_context:
        items = list(tool_context.state.get(STATE_ITEMS, []))

    item = {
        "id": len(items) + 1,
        "name": name,
        "unit_price": float(unit_price),
        "quantity": float(quantity),
        "unit": unit,
        "note": note,
    }
    items.append(item)

    if tool_context:
        tool_context.state[STATE_ITEMS] = items

    subtotal = unit_price * quantity
    return (
        f"✅ 品目を追加しました。\n"
        f"  [{item['id']}] {name}  {unit_price:,.0f}円 × {quantity}{unit} = {subtotal:,.0f}円\n"
        f"  現在の品目数: {len(items)}件"
    )

def remove_item(
    item_id: int,
    tool_context: ToolContext = None,
) -> str:
    """
    見積もりから品目をIDで削除する。

    Args:
        item_id: 削除する品目のID（list_items で確認できる番号）
    """
    if not tool_context:
        return "ToolContextがありません。"

    items: list[dict] = list(tool_context.state.get(STATE_ITEMS, []))
    before = len(items)
    items = [it for it in items if it["id"] != item_id]

    if len(items) == before:
        return f"⚠️ ID={item_id} の品目は見つかりませんでした。"

    # ID を振り直す
    for i, it in enumerate(items, 1):
        it["id"] = i

    tool_context.state[STATE_ITEMS] = items
    return f"🗑️ ID={item_id} の品目を削除しました。残り {len(items)} 件。"


def list_items(tool_context: ToolContext = None) -> str:
    """現在の見積もり品目一覧を表示する。"""
    if not tool_context:
        return "ToolContextがありません。"

    items: list[dict] = tool_context.state.get(STATE_ITEMS, [])
    if not items:
        return "⚠️ まだ品目が登録されていません。「品目を追加して」と話しかけてください。"

    lines = ["【現在の見積もり品目】\n"]
    for it in items:
        sub = it["unit_price"] * it["quantity"]
        note_str = f"  備考: {it['note']}" if it["note"] else ""
        lines.append(
            f"  [{it['id']}] {it['name']}\n"
            f"       単価: {it['unit_price']:,.0f}円 × {it['quantity']}{it['unit']}"
            f" = {sub:,.0f}円{note_str}"
        )
    return "\n".join(lines)


def calculate_quote(
    discount_rate: float = 0.0,
    tax_rate: float = 0.10,
    tool_context: ToolContext = None,
) -> str:
    """
    登録済み品目から見積もりを計算する。
    小計・値引き・消費税・合計を算出して state に保存する。
    計算結果は export_to_excel で Excel 出力に使われる。

    Args:
        discount_rate: 値引き率（0.0〜1.0 例: 0.1 = 10%引き）
        tax_rate: 消費税率（デフォルト 0.10 = 10%）
    """
    if not tool_context:
        return "ToolContextがありません。"

    items: list[dict] = tool_context.state.get(STATE_ITEMS, [])
    if not items:
        return "⚠️ 品目が登録されていません。先に品目を追加してください。"

    # ── 計算ロジック（Python で直接実行） ──
    rows = []
    subtotal_total = 0.0
    for it in items:
        sub = it["unit_price"] * it["quantity"]
        subtotal_total += sub
        rows.append({
            "id": it["id"],
            "name": it["name"],
            "unit_price": it["unit_price"],
            "quantity": it["quantity"],
            "unit": it["unit"],
            "subtotal": sub,
            "note": it.get("note", ""),
        })

    discount_amount = round(subtotal_total * discount_rate)
    subtotal_after_discount = subtotal_total - discount_amount
    tax_amount = round(subtotal_after_discount * tax_rate)
    total = subtotal_after_discount + tax_amount

    result = {
        "rows": rows,
        "subtotal": subtotal_total,
        "discount_rate": discount_rate,
        "discount_amount": discount_amount,
        "subtotal_after_discount": subtotal_after_discount,
        "tax_rate": tax_rate,
        "tax_amount": tax_amount,
        "total": total,
    }

    # state に保存
    tool_context.state[STATE_RESULT] = result

    # ── 返却テキスト（エージェントが表形式で表示する元データ） ──
    lines = ["【見積もり計算結果】\n"]
    lines.append(f"{'No.':<4} {'品目名':<20} {'単価':>12} {'数量':>6} {'単位':<6} {'小計':>14}")
    lines.append("─" * 68)
    for r in rows:
        lines.append(
            f"{r['id']:<4} {r['name']:<20} "
            f"{r['unit_price']:>12,.0f} {r['quantity']:>6.1f} {r['unit']:<6} {r['subtotal']:>14,.0f}"
        )
    lines.append("─" * 68)
    lines.append(f"{'小計':<44} {subtotal_total:>14,.0f} 円")
    if discount_rate > 0:
        lines.append(f"{'値引き(' + str(int(discount_rate*100)) + '%)':<44} {-discount_amount:>14,.0f} 円")
        lines.append(f"{'値引後小計':<44} {subtotal_after_discount:>14,.0f} 円")
    lines.append(f"{'消費税(' + str(int(tax_rate*100)) + '%)':<44} {tax_amount:>14,.0f} 円")
    lines.append("━" * 68)
    lines.append(f"{'合計金額':<44} {total:>14,.0f} 円")
    lines.append("\nExcelで出力する場合は「Excelで保存して」と言ってください。")

    return "\n".join(lines)


async def export_to_excel(
    filename: str = "見積もり.xlsx",
    title: str = "御見積書",
    company_name: str = "",
    tool_context: ToolContext = None,
) -> str:
    """
    計算済みの見積もりを Excel ファイルとして ADK Artifact に保存する。
    事前に calculate_quote を実行しておく必要がある。

    Args:
        filename: ファイル名（デフォルト: 見積もり.xlsx）
        title: 見積書タイトル（デフォルト: 御見積書）
        company_name: 宛先会社名（オプション）
    """
    if not tool_context:
        return "ToolContextがありません。"

    result: dict = tool_context.state.get(STATE_RESULT)
    if not result:
        # フォールバック: items から直接計算
        items = tool_context.state.get(STATE_ITEMS, [])
        if not items:
            return "⚠️ 見積もりデータがありません。先に品目を追加・計算してください。"
        rows = []
        subtotal = 0
        for it in items:
            sub = it["unit_price"] * it["quantity"]
            subtotal += sub
            rows.append({**it, "subtotal": sub})
        tax = round(subtotal * 0.10)
        result = {
            "rows": rows,
            "subtotal": subtotal,
            "discount_rate": 0,
            "discount_amount": 0,
            "subtotal_after_discount": subtotal,
            "tax_rate": 0.10,
            "tax_amount": tax,
            "total": subtotal + tax,
        }

    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    wb = _build_workbook(result, title, company_name)
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    if tool_context:
        try:
            artifact = types.Part.from_bytes(
                data=excel_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            version = await tool_context.save_artifact(filename=filename, artifact=artifact)
            rows_count = len(result.get("rows", []))
            total = result.get("total", 0)
            return (
                f"✅ '{filename}' を保存しました（version {version}）。\n"
                f"   品目数: {rows_count}件 / 合計金額: {total:,.0f}円\n"
                f"ADK UIの📎アイコンからダウンロードできます。"
            )
        except Exception as e:
            pass

    # フォールバック: デスクトップに保存
    save_path = os.path.join(os.path.expanduser("~/Desktop"), filename)
    with open(save_path, "wb") as f:
        f.write(excel_bytes)
    return f"✅ デスクトップに '{filename}' として保存しました。"


def reset_quote(tool_context: ToolContext = None) -> str:
    """見積もりをリセットして最初からやり直す。"""
    if not tool_context:
        return "ToolContextがありません。"
    tool_context.state[STATE_ITEMS] = []
    tool_context.state[STATE_RESULT] = None
    tool_context.state["quote:pending_code"] = None
    return "🔄 見積もりをリセットしました。新しい品目を追加してください。"


# ────────────────────────────────────────────
#  Excel ワークブック生成（内部関数）
# ────────────────────────────────────────────

def _build_workbook(result: dict, title: str, company_name: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "見積もり"

    # ── スタイル定義 ──
    HEADER_BG = "4472C4"
    SUBTOTAL_BG = "D9E1F2"
    TOTAL_BG = "1F3864"

    def h_style(cell, bg=HEADER_BG, color="FFFFFF", bold=True):
        cell.font = Font(bold=bold, color=color)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def border(cell):
        cell.border = thin

    # ── タイトル行 ──
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 宛先
    row = 2
    if company_name:
        ws[f"A{row}"] = f"宛先: {company_name} 御中"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

    row += 1  # 空行

    # ── ヘッダー行 ──
    headers = ["No.", "品目名", "単価（円）", "数量", "単位", "小計（円）", "備考"]
    col_widths = [6, 30, 14, 8, 8, 16, 20]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col, value=h)
        h_style(cell)
        border(cell)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 20
    header_row = row
    row += 1

    # ── 品目行 ──
    rows_data = result.get("rows", [])
    for it in rows_data:
        data = [
            it["id"], it["name"],
            it["unit_price"], it["quantity"], it["unit"],
            it["subtotal"], it.get("note", ""),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            border(cell)
            if col in (3, 6):  # 金額列
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # ── 集計行 ──
    def summary_row(label: str, amount: float, bg=SUBTOTAL_BG, bold=False, color="000000"):
        ws.merge_cells(f"A{row}:E{row}")
        lc = ws[f"A{row}"]
        lc.value = label
        lc.font = Font(bold=bold, color=color)
        lc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border = thin

        ac = ws[f"F{row}"]
        ac.value = amount
        ac.font = Font(bold=bold, color=color)
        ac.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ac.alignment = Alignment(horizontal="right", vertical="center")
        ac.number_format = '#,##0'
        ac.border = thin

        ws[f"G{row}"].border = thin

    row += 1  # 空行
    summary_row("小計", result["subtotal"])
    row += 1

    dr = result["discount_rate"]
    if dr > 0:
        label = f"値引き（{int(dr*100)}%）"
        summary_row(label, -result["discount_amount"])
        row += 1
        summary_row("値引後小計", result["subtotal_after_discount"])
        row += 1

    tr = result["tax_rate"]
    summary_row(f"消費税（{int(tr*100)}%）", result["tax_amount"])
    row += 1

    # 合計
    ws.merge_cells(f"A{row}:E{row}")
    lc = ws[f"A{row}"]
    lc.value = "合計金額"
    lc.font = Font(bold=True, color="FFFFFF", size=12)
    lc.fill = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
    lc.alignment = Alignment(horizontal="right", vertical="center")
    lc.border = thin
    ws.row_dimensions[row].height = 22

    ac = ws[f"F{row}"]
    ac.value = result["total"]
    ac.font = Font(bold=True, color="FFFFFF", size=12)
    ac.fill = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
    ac.alignment = Alignment(horizontal="right", vertical="center")
    ac.number_format = '#,##0'
    ac.border = thin
    ws[f"G{row}"].fill = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
    ws[f"G{row}"].border = thin

    # オートフィルター（品目エリア）
    ws.auto_filter.ref = f"A{header_row}:G{header_row + len(rows_data)}"

    return wb


# ────────────────────────────────────────────
#  ADK FunctionTool 登録
# ────────────────────────────────────────────
quote_add_tool = FunctionTool(add_item)
quote_remove_tool = FunctionTool(remove_item)
quote_list_tool = FunctionTool(list_items)
quote_calculate_tool = FunctionTool(calculate_quote)
quote_excel_tool = FunctionTool(export_to_excel)
quote_reset_tool = FunctionTool(reset_quote)
